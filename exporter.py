"""
exporter.py — v3
2 mode export:
1. export_weekly  — weekly analysis report
2. export_restock — restock recommendations report
"""
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

C_NAVY = "1E3A5F"; C_BLUE = "2E75B6"; C_RED = "CC0000"
C_AMBER = "886600"; C_GREEN = "15803D"; C_ALT = "F5F5F5"
C_WHITE = "FFFFFF"
TIER_BG = {
    "1. Best Seller": "D1FAE5", "2. Uprising": "DBEAFE",
    "3. Slow Moving": "FEF3C7", "4. Deadweight": "F3F4F6", "5. Sin": "FFE4E6",
}
TIER_FG = {
    "1. Best Seller": "15803D", "2. Uprising": "1D4ED8",
    "3. Slow Moving": "92400E", "4. Deadweight": "6B7280", "5. Sin": "9F1239",
}


def _b():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, row, ncols, bg=C_NAVY):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color=C_WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _b()


def _title(ws, text, ncols):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=13, color=C_NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


def export_weekly(weekly_result: dict, metrics: dict, output_path: str = None) -> str:
    """Export weekly analysis report ke Excel."""
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = f"output/weekly_report_{ts}.xlsx"
    os.makedirs("output", exist_ok=True)

    wb = openpyxl.Workbook()
    ring      = metrics.get("ringkasan", {})
    tier_data = metrics.get("tier_data")
    semua     = metrics.get("semua_produk")

    # ── Sheet 1: Weekly Summary ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Weekly Summary"
    _title(ws, f"WEEKLY PURCHASING REPORT — {datetime.now().strftime('%d %B %Y')}", 6)

    ws.merge_cells("A2:F2")
    ws["A2"] = weekly_result.get("ringkasan_minggu", "")
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    # Stats
    stats = [
        ("Total SKU", ring.get("total_produk", 0)),
        ("Perlu di-PO", ring.get("perlu_po", 0)),
        ("URGENT", ring.get("urgent", 0)),
        ("NORMAL", ring.get("normal", 0)),
        ("Stok Aman", ring.get("stok_cukup", 0)),
        ("Skip (PO aktif)", ring.get("skip_po_berjalan", 0)),
    ]
    for ci, (label, val) in enumerate(stats, 1):
        ws.cell(row=4, column=ci, value=label).font = Font(bold=True, size=10)
        cell = ws.cell(row=5, column=ci, value=val)
        cell.font = Font(bold=True, size=16, color=C_NAVY)
        cell.alignment = Alignment(horizontal="center")

    # Alert Best Seller
    row = 7
    ws.cell(row=row, column=1, value="🚨 BEST SELLER STOK KRITIS:").font = Font(bold=True, color=C_RED)
    for item in weekly_result.get("alert_best_seller", []):
        row += 1
        ws.cell(row=row, column=1, value=f"  • {item}")

    row += 2
    ws.cell(row=row, column=1, value="✅ HIGHLIGHT POSITIF:").font = Font(bold=True, color=C_GREEN)
    for item in weekly_result.get("highlight_positif", []):
        row += 1
        ws.cell(row=row, column=1, value=f"  • {item}")

    row += 2
    ws.cell(row=row, column=1, value="⚠️ PERLU DIWASPADAI:").font = Font(bold=True, color=C_AMBER)
    for item in weekly_result.get("highlight_negatif", []):
        row += 1
        ws.cell(row=row, column=1, value=f"  • {item}")

    row += 2
    ws.cell(row=row, column=1, value="📋 REKOMENDASI TINDAKAN:").font = Font(bold=True, color=C_NAVY)
    for item in weekly_result.get("rekomendasi_tindakan", []):
        row += 1
        ws.cell(row=row, column=1, value=f"  • {item}")

    row += 2
    ws.cell(row=row, column=1, value=f"💡 SARAN UTAMA: {weekly_result.get('saran_utama','')}").font = Font(bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    for ci, w in enumerate([35, 20, 15, 15, 15, 15], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Performa Produk ─────────────────────────────────────────
    if tier_data is not None and not tier_data.empty:
        ws2 = wb.create_sheet("Performa Produk")
        prod_col = "Product" if "Product" in tier_data.columns else tier_data.columns[0]
        headers  = [prod_col, "Avg Mingguan", "Total Terjual", "Periode Aktif", "Tier", "Tren", "% Perubahan"]
        show_cols = [c for c in headers if c in tier_data.columns or c == prod_col]

        for ci, h in enumerate(show_cols, 1):
            ws2.cell(row=1, column=ci, value=h)
        _hdr(ws2, 1, len(show_cols), bg=C_BLUE)

        for ri, (_, row_data) in enumerate(tier_data.iterrows(), 2):
            tier = str(row_data.get("Tier", "-"))
            bg   = TIER_BG.get(tier, C_ALT)
            fg   = TIER_FG.get(tier, "333333")
            for ci, col in enumerate(show_cols, 1):
                val  = row_data.get(col, "")
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.border = _b()
                cell.alignment = Alignment(vertical="center")
                if col == "Tier":
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(bold=True, color=fg)
                elif ri % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=C_ALT)

        for ci, w in enumerate([40, 15, 15, 15, 20, 12, 14], 1):
            ws2.column_dimensions[get_column_letter(ci)].width = w
        ws2.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\n💾 Weekly report: {output_path}")
    return output_path


def export_restock(restock_result: dict, metrics: dict, output_path: str = None) -> str:
    """Export restock recommendations ke Excel."""
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = f"output/restock_{ts}.xlsx"
    os.makedirs("output", exist_ok=True)

    wb   = openpyxl.Workbook()
    ring = metrics.get("ringkasan", {})

    # ── Sheet 1: Rekomendasi PO ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Rekomendasi PO"
    _title(ws, f"RESTOCK RECOMMENDATIONS — {datetime.now().strftime('%d %B %Y %H:%M')}", 9)

    ws.merge_cells("A2:I2")
    ws["A2"] = restock_result.get("analisis_singkat", "")
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws["A2"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[2].height = 18

    headers = ["No", "Nama Produk", "Tier", "Tren", "Coverage", "Prioritas", "Qty Order", "Est. Tiba", "Alasan"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    _hdr(ws, 4, len(headers))
    ws.row_dimensions[4].height = 26

    reko = restock_result.get("rekomendasi_po", [])
    for i, item in enumerate(reko, 1):
        row  = 4 + i
        prio = item.get("prioritas", "NORMAL")
        tier = item.get("tier", "-")
        bg   = "FFE2E2" if prio == "URGENT" else ("FFF9E2" if i % 2 == 0 else C_WHITE)

        vals = [
            i, item.get("nama_produk", ""), tier, item.get("tren", "-"),
            item.get("coverage_sekarang", "-"), prio,
            item.get("qty_rekomendasi", 0), item.get("estimasi_tiba", "-"),
            item.get("alasan", "-"),
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = _b()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if ci == 3:
                cell.fill = PatternFill("solid", fgColor=TIER_BG.get(tier, C_ALT))
                cell.font = Font(bold=True, color=TIER_FG.get(tier, "333333"), size=10)
            if ci == 6:
                cell.font = Font(bold=True, color=C_RED if prio == "URGENT" else C_AMBER)
        ws.row_dimensions[row].height = 22

    for ci, w in enumerate([5, 40, 18, 12, 14, 12, 12, 22, 45], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A5"

    # ── Sheet 2: Ringkasan ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Ringkasan")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 22

    ws2.merge_cells("A1:B1")
    ws2["A1"] = "RINGKASAN RESTOCK"
    ws2["A1"].font = Font(bold=True, size=13, color=C_NAVY)
    ws2["A1"].alignment = Alignment(horizontal="center")

    rows_data = [
        ("Tanggal Run",           datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Total SKU",             ring.get("total_produk", 0)),
        ("Perlu di-PO",           ring.get("perlu_po", 0)),
        ("URGENT",                ring.get("urgent", 0)),
        ("NORMAL",                ring.get("normal", 0)),
        ("Skip (PO Berjalan)",    ring.get("skip_po_berjalan", 0)),
        ("Stok Cukup",            ring.get("stok_cukup", 0)),
        ("Saran Tindakan",        restock_result.get("saran_tindakan", "")),
    ]
    for ri, (label, val) in enumerate(rows_data, 3):
        cl = ws2.cell(row=ri, column=1, value=label)
        cv = ws2.cell(row=ri, column=2, value=val)
        cl.font = Font(bold=True)
        cl.border = _b()
        cv.border = _b()
        if ri % 2 == 0:
            cl.fill = cv.fill = PatternFill("solid", fgColor=C_ALT)

    wb.save(output_path)
    print(f"\n💾 Restock report: {output_path}")
    return output_path
